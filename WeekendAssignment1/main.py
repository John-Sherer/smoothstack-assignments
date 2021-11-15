import logging
from openpyxl import load_workbook
import sys
import datetime

if __name__ == '__main__':
    # Used to confirm that month in the input is valid
    months = ['january', 'february', 'march', 'april', 'may', 'june', 'july', 'august', 'september', 'october',
              'november', 'december']

    # Load the logger
    logging.basicConfig(filename='logfile.log', filemode='w', level=logging.DEBUG)
    month_name = ""
    month_name_capitalized = ""  # For prettier logging
    file_name = ""

    # Check input is valid
    if len(sys.argv) < 2:
        # Missing arguments
        print('Error, please provide the month file name as input')
        exit(0)
    else:
        str_input = sys.argv[1]
        if str_input[-6:] != '.xlsx)':
            # Bad file type
            print('Error, input is not an .xlsx file, or in the incorrect format.' +
                  'Input should be of the form "month(filename.xlsx)", without the quotation marks.')
            exit(1)
        if str_input.find('(') == -1:
            # Missing paren
            print('Error, no matching parenthesis found. Input should be of the form "month(filename.xlsx)".')
            exit(1)
        month_name = str_input[:str_input.index('(')].lower()
        month_name_capitalized = month_name[0].upper() + month_name[1:]
        if month_name not in months:
            # Month DNE
            print('Error, could not find month named "{}"'.format(str_input[:str_input.index('(')]))
            exit(1)
        file_name = str_input[str_input.index('(') + 1: str_input.index(')')]

    logging.info('Input validated, file = "{}", month = "{}".'.format(file_name, month_name))
    workbook = load_workbook(file_name)

    # Verify sheet names are valid
    sheet_names = workbook.sheetnames
    if sheet_names != ['Summary Rolling MoM', 'VOC Rolling MoM', 'Monthly Verbatim Statements']:
        logging.warning("Error, could not verity correct sheet names. Should be 'Summary Rolling MoM'," +
                        "'VOC Rolling MoM', and 'Monthly Verbatim Statements'.")
        exit(1)
    logging.info('Sheet names validated')

    # Start with summary
    summary_sheet = workbook['Summary Rolling MoM']

    # Validate column titles and store their indices
    expected_column_titles = ['Calls Offered', 'Abandon after 30s', 'FCR', 'DSAT', 'CSAT']
    expected_column_locations = [-1, -1, -1, -1, -1]
    for i in range(0, 5):
        found_title = False
        for cel in summary_sheet[1]:
            if expected_column_titles[i] == str(cel.value).strip(' '):
                found_title = True
                expected_column_locations[i] = summary_sheet[1].index(cel)
        if not found_title:
            logging.warning("Warning: could not find a column for " + expected_column_titles[i])
            exit(1)

    # Log info
    for row in summary_sheet.values:
        if type(row[0]) is datetime.datetime and row[0].date().month == months.index(month_name) + 1:
            logging.info("Data for " + month_name_capitalized + ":")
            logging.info("\tCalls Offered: {}".format(row[expected_column_locations[0]]))
            logging.info("\tAbandon after 30s: %3.2f" % (row[expected_column_locations[1]] * 100) + "%")
            logging.info("\tFCR: %3.2f" % (row[expected_column_locations[2]] * 100) + "%")
            logging.info("\tDSAT: %3.2f" % (row[expected_column_locations[3]] * 100) + "%")
            logging.info("\tCSAT: %3.2f" % (row[expected_column_locations[4]] * 100) + "%")

    # Continue with VOC sheet
    voc_sheet = workbook['VOC Rolling MoM']

    # Find the column with the first instance of the desired month
    col = -1
    for i in range(0, len(voc_sheet[1])):
        cel = voc_sheet[1][i]
        if type(cel.value) is str and cel.value[:3].lower() == month_name[:3]:
            col = i
            break
        elif type(cel.value) is datetime.datetime and cel.value.date().month == months.index(month_name) + 1:
            col = i
            break

    # Find the rows each of the desired fields: Promoters, Passives, and Detractors
    promoters = -1
    passives = -1
    detractors = -1
    for row in voc_sheet.values:
        if type(row[0]) is str:
            if row[0][:9] == 'Promoters':
                promoters = row[col]
            elif row[0][:8] == 'Passives':
                passives = row[col]
            elif row[0][:10] == 'Detractors' or row[0][:11] == 'Dectractors':
                detractors = row[col]

    # Verify that we have found values
    if promoters == -1 or passives == -1 or detractors == -1:
        logging.warning("Warning, could not find one or more of promoters, passives and detractors.")
        exit(1)

    # Print to log
    if promoters >= 200:
        logging.info('\tPromoters: Good ({})'.format(promoters))
    else:
        logging.info('\tPromoters: Bad ({})'.format(promoters))
    if passives >= 100:
        logging.info('\tPassives: Good: ({})'.format(passives))
    else:
        logging.info('\tPassives: Bad ({})'.format(passives))
    if detractors >= 100:
        logging.info('\tDetractors: Good ({})'.format(detractors))
    else:
        logging.info('\tDetractors: Bad ({})'.format(detractors))
